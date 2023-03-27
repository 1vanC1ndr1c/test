import datetime
import os
import random
import sys
import typing
from pathlib import Path

import openpyxl


def main():
    downloads_dir = Path.home() / 'Downloads'

    era_file_path = [str(era) for era in os.listdir(downloads_dir)
                     if str(era).startswith('ERA')][0]
    era_file_path = downloads_dir / era_file_path

    modify_file(era_file_path)

    rename_file(era_file_path, downloads_dir)


def rename_file(era_file_path, downloads_dir):
    new_file_name = f'ERA-{month()}-CindrićIvan.{era_file_path.suffix}'
    os.rename(era_file_path,
              downloads_dir / new_file_name)


def modify_file(era_file_path):
    era_file = openpyxl.load_workbook(era_file_path)

    lookup_sheet = era_file['Lookup']
    days_indices = get_days_indices(lookup_sheet)

    no_of_lines = len(days_indices)

    work_sheet = era_file['Evidencija']
    paste_names(work_sheet, no_of_lines)
    paste_dates(days_indices, lookup_sheet, work_sheet)
    paste_work_hours(work_sheet, no_of_lines)
    paste_projects(work_sheet, no_of_lines)

    era_file.save(era_file_path)


def get_days_indices(sheet) -> typing.List[typing.Tuple[int, int]]:
    days_index = [index
                  for index, column_cell in enumerate(sheet.iter_rows())
                  if str(column_cell[0].value).startswith('Radni dani')][0]
    days_index += 2

    indices = []
    for index, column_cell in enumerate(sheet.iter_rows(days_index)):
        if not column_cell[0].value:
            break
        indices.append((days_index + index, 1))

    return indices


def paste_names(work_sheet, no_of_lines):
    _paste(work_sheet, no_of_lines, 'Prezime', 'Cindrić, Ivan')


def paste_dates(days_indices, lookup_sheet, work_sheet):
    row_index, column_index = _find_pos(work_sheet, 'Dan')
    for i, day_index in enumerate(days_indices):
        day_row, day_column = day_index
        day = lookup_sheet.cell(row=day_row, column=day_column).value
        work_sheet.cell(row=row_index + i, column=column_index).value = day


def paste_projects(work_sheet, no_of_lines):
    # 'Istraživanje # 90-20-00004/850 # BEYOND # EU',
    # 'Komercijalni # 25-22-00002-850 # Proza Hat IEC 60870-5-104 Secure gateway # KET' # NOQA
    # 'Imovina # 10-22-00001/850 # Proza HAT EDS # n/a'
    print('How many projects? = ', end='')
    no_of_projects = int(input())

    projects = []
    for no in range(1, no_of_projects + 1):
        print(f'Project {no} = ', end='')
        project = input()

        print(f'Percentage of days (%) = ', end='')
        percentage = float(input())

        projects.append({'project': project, 'percentage': percentage})

    if sum(x['percentage'] for x in projects) != 100:
        print('Wrong percentages, randomizing...')
        percentages = add_to_hundred(len(projects))
        for index, pr in enumerate(projects):
            pr['percentage'] = percentages[index]
            print(f'{pr["project"]} = {pr["percentage"]}%')

    row_index, column_index = _find_pos(work_sheet, 'Tip')
    start = 0

    for index, project in enumerate(projects):

        if index == len(projects) - 1:
            end = no_of_lines
        else:
            end = start + int(no_of_lines * (project['percentage'] / 100))

        for i in range(start, end):
            work_sheet.cell(row=row_index + i,
                            column=column_index).value = project['project']
        start = end
        print(f'{end = }, {no_of_lines = }')


def paste_work_hours(work_sheet, no_of_lines):
    _paste(work_sheet, no_of_lines, 'Sati', 8)


def _paste(sheet, no_of_lines, column_value, paste_value):
    row_index, column_index = _find_pos(sheet, column_value)

    for i in range(no_of_lines):
        sheet.cell(row=row_index + i, column=column_index).value = paste_value


def _find_pos(sheet, column_value):
    for row_index, row in enumerate(sheet.iter_rows()):
        for column_index, column in enumerate(row[:5]):
            if str(column.value).startswith(str(column_value)):
                if column_index > 10:
                    continue
                return row_index + 2, column_index + 1


def add_to_hundred(no_of_el):
    dividers = sorted(random.sample(range(1, 100), no_of_el - 1))
    return [a - b for a, b in zip(dividers + [100], [0] + dividers)]


def month() -> str:
    today_month = datetime.datetime.now().month
    return ('_', 'Siječanj', 'Veljača', 'Ožujak', 'Travanj', 'Svibanj',
            'Lipanj', 'Srpanj', 'Kolovoz', 'Rujan', 'Listopa', 'Studeni',
            'Prosinac')[today_month]


if __name__ == '__main__':
    sys.exit(main())
